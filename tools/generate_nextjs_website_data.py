from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
PROJECTS_ROOT = ROOT / "projects"
OUTPUTS_ROOT = ROOT / "11-outputs"
WEBSITE_PUBLIC = ROOT / "website" / "public"
DATA_ROOT = WEBSITE_PUBLIC / "data"
GENERATED_ROOT = WEBSITE_PUBLIC / "generated"


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-")
    return value or "project"


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "n/a", "na", "-"}:
        return None
    text = text.replace(",", "").replace("%", "").replace("EGP", "").strip()
    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return number


def safe_percent(value: Any) -> float | None:
    number = safe_float(value)
    if number is None:
        return None
    if number > 1:
        number = number / 100.0
    return max(0.0, min(number, 10.0))


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "cp1252"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except Exception:
            continue
    return []


def compact_file_record(path: Path, base: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "relative_path": path.relative_to(base).as_posix(),
        "extension": path.suffix.lower().lstrip(".") or "file",
        "size_kb": round(stat.st_size / 1024, 1),
        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
    }


def list_project_files(path: Path, base: Path, limit: int = 120) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    files = [
        compact_file_record(child, base)
        for child in sorted(path.rglob("*"))
        if child.is_file() and not child.name.startswith(".")
    ]
    return files[:limit]


def preview_table(path: Path, limit: int = 8) -> dict[str, Any]:
    rows = read_csv_rows(path)
    columns: list[str] = []
    if rows:
        columns = list(rows[0].keys())
    elif path.exists():
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                columns = next(reader, [])
        except Exception:
            columns = []
    return {
        "file": path.name,
        "exists": path.exists(),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": columns,
        "rows": rows[:limit],
    }


def xlsx_summary(path: Path, limit: int = 8) -> dict[str, Any]:
    summary = {"file": path.name, "exists": path.exists(), "sheets": []}
    if not path.exists():
        return summary
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter, [])]
            preview_rows = []
            row_count = 0
            for row in rows_iter:
                row_count += 1
                if len(preview_rows) < limit:
                    preview_rows.append({headers[idx] if idx < len(headers) and headers[idx] else f"Column {idx + 1}": value for idx, value in enumerate(row)})
            summary["sheets"].append(
                {
                    "name": sheet_name,
                    "row_count": row_count,
                    "column_count": len(headers),
                    "columns": headers,
                    "rows": preview_rows,
                }
            )
    except Exception as exc:
        summary["error"] = str(exc)
    return summary


def sqlite_table_counts(path: Path) -> dict[str, Any]:
    result: dict[str, Any] = {"exists": path.exists(), "tables": {}, "error": None}
    if not path.exists():
        return result
    try:
        connection = sqlite3.connect(path)
        cursor = connection.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        for (table_name,) in cursor.fetchall():
            try:
                count = connection.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            except Exception:
                count = None
            result["tables"][table_name] = count
        connection.close()
    except Exception as exc:
        result["error"] = str(exc)
    return result


def build_feature_payload(project: dict[str, Any], rows: dict[str, list[dict[str, str]]]) -> dict[str, Any]:
    base = Path(project["path"])
    data_dir = base / "01-data" / "import_templates"
    delay_dir = base / "02-delay_analysis" / "steel_delay_tia_templates"
    schedule_dir = base / "03-schedule"
    contracts_dir = base / "05-contracts"
    evidence_dir = base / "06-evidence"
    letters_dir = base / "07-letters_intelligence"
    outputs_dir = OUTPUTS_ROOT / project["project_folder_name"]

    delay_templates = [
        preview_table(path)
        for path in sorted(delay_dir.glob("*.csv"))
    ]
    delay_required_names = [
        "01-project_metadata_template.csv",
        "02-master_activity_steel_analysis.csv",
        "03-employer_steel_supply_at_site.csv",
        "04-p6_activity_export.csv",
        "05-relationship_file.csv",
        "06-contract_library.csv",
        "07-ifc_conflict.csv",
        "08-payments.csv",
        "09-rfi_status.csv",
        "10-contractor_steel_supplied_at_site.csv",
        "11-concurrency_matrix_template.updated.csv",
    ]
    normalized_delay_files = {re.sub(r"\s+", "", item["file"].lower()): item for item in delay_templates}
    missing_delay = [
        name for name in delay_required_names
        if re.sub(r"\s+", "", name.lower()) not in normalized_delay_files
    ]

    letter_files = list_project_files(letters_dir / "inbox", base, 160)
    letter_workbook = xlsx_summary(letters_dir / "letters_intelligence.xlsx")
    contract_files = list_project_files(contracts_dir / "source", base, 80)
    evidence_files = list_project_files(evidence_dir, base, 80)
    output_files = list_project_files(outputs_dir, OUTPUTS_ROOT, 20)

    return {
        "overview": {
            "data_sources": {key: len(value) for key, value in rows.items()},
            "source_tables": {
                "projects": preview_table(data_dir / "projects.csv"),
                "activities": preview_table(data_dir / "activities.csv"),
                "progress_updates": preview_table(data_dir / "progress_updates.csv"),
                "evm": preview_table(data_dir / "evm.csv"),
                "risks": preview_table(data_dir / "risks.csv"),
                "claims": preview_table(data_dir / "claims.csv"),
                "contracts": preview_table(data_dir / "contracts.csv"),
                "payments": preview_table(data_dir / "payments.csv"),
                "milestones": preview_table(data_dir / "milestones.csv"),
                "delay_events": preview_table(data_dir / "delay_events.csv"),
                "wbs": preview_table(data_dir / "wbs.csv"),
                "s_curve": preview_table(data_dir / "s_curve.csv"),
            },
        },
        "letters_intelligence": {
            "folder": "07-letters_intelligence",
            "inbox_files": letter_files,
            "inbox_file_count": len(letter_files),
            "workbook": letter_workbook,
            "detectors": [
                {"name": "Inbox folder detector", "status": "Active" if (letters_dir / "inbox").exists() else "Missing", "detail": "Recognizes new PDF, DOCX, XLSX, CSV, and message files under project letters inbox."},
                {"name": "Letters workbook detector", "status": "Active" if (letters_dir / "letters_intelligence.xlsx").exists() else "Missing", "detail": "Reads the project-specific letters intelligence workbook when available."},
                {"name": "Project isolation", "status": "Active", "detail": "Only files inside the selected project folder are listed."},
            ],
        },
        "delay_analysis": {
            "folder": "02-delay_analysis/steel_delay_tia_templates",
            "templates": delay_templates,
            "required_file_count": len(delay_required_names),
            "recognized_file_count": len(delay_templates),
            "missing_required_files": missing_delay,
            "schedule_tables": {
                "MEP Activities": preview_table(schedule_dir / "MEP Activities.csv"),
                "MEP Schedule": preview_table(schedule_dir / "MEP Schedule.csv"),
                "MEP Civil Logic": preview_table(schedule_dir / "MEP Civil Logic.csv"),
                "BL Schedule": preview_table(schedule_dir / "BL Schedule.csv"),
            },
            "detectors": [
                {"name": "Delay TIA template detector", "status": "Ready" if not missing_delay else "Needs files", "detail": f"{len(delay_templates)} CSV files recognized in the selected project."},
                {"name": "Column inspector", "status": "Active", "detail": "Every detected CSV includes row count, column count, and preview rows."},
                {"name": "MEP schedule detector", "status": "Active" if (schedule_dir / "MEP Schedule.csv").exists() else "Missing", "detail": "Recognizes project-specific MEP schedule and civil logic tables."},
            ],
        },
        "contract_claims": {
            "folder": "05-contracts",
            "source_files": contract_files,
            "evidence_files": evidence_files,
            "database": sqlite_table_counts(contracts_dir / "contract_claims.db"),
            "clause_library": xlsx_summary(contracts_dir / "source" / "Overall_Contract_clause_library.xlsx"),
            "detectors": [
                {"name": "Contract source detector", "status": "Active" if contract_files else "Missing", "detail": "Finds contract PDFs and clause libraries inside the selected project only."},
                {"name": "Knowledge base detector", "status": "Active" if (contracts_dir / "contract_claims.db").exists() else "Missing", "detail": "Uses the selected project's own SQLite knowledge base."},
                {"name": "Evidence detector", "status": "Active" if evidence_files else "Missing", "detail": "Maps claim evidence files from the selected project evidence folder."},
            ],
        },
        "outputs_and_watchers": {
            "outputs_folder": f"11-outputs/{project['project_folder_name']}",
            "output_files": output_files,
            "watchers": [
                {"name": "Project data detector", "status": "Active", "detail": "Website data generator fingerprints every selected project folder."},
                {"name": "No-Git sync watcher", "status": "Configured" if (ROOT / "RUN_FULL_PROJECT_NO_GIT_SYNC.bat").exists() else "Missing", "detail": "Syncs local code, project folders, generated HTML, and website files to GitHub."},
                {"name": "Generated HTML output watcher", "status": "Configured" if outputs_dir.exists() else "Missing", "detail": "Publishes project-specific HTML outputs from 11-outputs."},
            ],
        },
    }


def pick(row: dict[str, Any], names: list[str]) -> Any:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        key = name.strip().lower()
        if key in lowered and str(lowered[key]).strip() != "":
            return lowered[key]
    return None


def first_valid(rows: list[dict[str, Any]], names: list[str]) -> Any:
    for row in rows:
        value = pick(row, names)
        if value not in (None, ""):
            return value
    return None


def sum_column(rows: list[dict[str, Any]], names: list[str]) -> float | None:
    total = 0.0
    seen = False
    for row in rows:
        value = safe_float(pick(row, names))
        if value is not None:
            total += value
            seen = True
    return total if seen else None


def average(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    return sum(clean) / len(clean) if clean else None


def latest_mtime(path: Path) -> str | None:
    if not path.exists():
        return None
    latest = 0.0
    for child in path.rglob("*"):
        if child.is_file():
            latest = max(latest, child.stat().st_mtime)
    if latest == 0:
        return None
    return datetime.fromtimestamp(latest).isoformat(timespec="seconds")


def fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    if not path.exists():
        return ""
    for child in sorted(p for p in path.rglob("*") if p.is_file()):
        rel = child.relative_to(path).as_posix()
        if any(part in {".git", ".venv", "node_modules", "__pycache__"} for part in child.parts):
            continue
        digest.update(rel.encode("utf-8"))
        digest.update(str(child.stat().st_size).encode("ascii"))
        digest.update(str(int(child.stat().st_mtime)).encode("ascii"))
    return digest.hexdigest()


def discover_projects() -> list[dict[str, Any]]:
    projects: list[dict[str, Any]] = []
    if not PROJECTS_ROOT.exists():
        return projects
    for sector_dir in sorted(p for p in PROJECTS_ROOT.iterdir() if p.is_dir() and not p.name.startswith("_")):
        for project_dir in sorted(p for p in sector_dir.iterdir() if p.is_dir() and not p.name.startswith("_")):
            manifest = read_json(project_dir / "project_manifest.json")
            project_json = read_json(project_dir / "project.json")
            project_id = manifest.get("project_id") or project_json.get("project_id") or slugify(project_dir.name).lower()
            display_name = (
                manifest.get("project_display_name")
                or project_json.get("project_display_name")
                or project_json.get("name")
                or project_dir.name
            )
            projects.append(
                {
                    "project_id": str(project_id),
                    "project_key": slugify(str(project_id)).lower(),
                    "project_folder_name": project_dir.name,
                    "project_display_name": str(display_name),
                    "sector": sector_dir.name,
                    "path": project_dir,
                }
            )
    return projects


def build_project_record(project: dict[str, Any]) -> dict[str, Any]:
    base = Path(project["path"])
    data_dir = base / "01-data" / "import_templates"
    rows = {
        "projects": read_csv_rows(data_dir / "projects.csv"),
        "contracts": read_csv_rows(data_dir / "contracts.csv"),
        "payments": read_csv_rows(data_dir / "payments.csv"),
        "progress": read_csv_rows(data_dir / "progress_updates.csv"),
        "evm": read_csv_rows(data_dir / "evm.csv"),
        "risks": read_csv_rows(data_dir / "risks.csv"),
        "claims": read_csv_rows(data_dir / "claims.csv"),
        "activities": read_csv_rows(data_dir / "activities.csv"),
        "milestones": read_csv_rows(data_dir / "milestones.csv"),
        "delay_events": read_csv_rows(data_dir / "delay_events.csv"),
    }

    project_meta = rows["projects"][0] if rows["projects"] else {}
    contract_value = (
        safe_float(first_valid(rows["projects"], ["contract_value", "budget", "bac"]))
        or sum_column(rows["contracts"], ["contract_value", "value", "amount", "bac"])
        or safe_float(first_valid(rows["evm"], ["bac", "budget_at_completion"]))
    )
    paid_amount = sum_column(rows["payments"], ["paid_amount", "paid", "payment_amount", "amount"])
    spent_amount = (
        sum_column(rows["payments"], ["actual_cost", "spent_amount", "cost", "ac"])
        or safe_float(first_valid(rows["evm"], ["ac", "actual_cost"]))
        or paid_amount
    )
    planned_progress = (
        safe_percent(first_valid(rows["progress"], ["planned_progress", "planned %", "planned_percent"]))
        or safe_percent(first_valid(rows["evm"], ["planned_progress", "planned_percent"]))
    )
    actual_progress = (
        safe_percent(first_valid(rows["progress"], ["actual_progress", "progress", "actual %", "actual_percent"]))
        or safe_percent(first_valid(rows["projects"], ["progress", "actual_progress"]))
    )

    bac = contract_value
    pv = bac * planned_progress if bac is not None and planned_progress is not None else None
    ev = bac * actual_progress if bac is not None and actual_progress is not None else None
    ac = spent_amount
    spi = ev / pv if ev is not None and pv and pv > 0 else safe_float(first_valid(rows["evm"], ["spi"]))
    cpi = ev / ac if ev is not None and ac and ac > 0 else safe_float(first_valid(rows["evm"], ["cpi"]))
    remaining_value = contract_value - (paid_amount or spent_amount or 0) if contract_value is not None else None
    risk_values = [safe_float(pick(row, ["risk_score", "score", "severity"])) for row in rows["risks"]]
    risk_score = average(risk_values)
    high_risk_count = sum(1 for value in risk_values if value is not None and value >= 70)
    delay_days = sum_column(rows["delay_events"], ["delay_days", "duration", "delay_duration"]) or 0
    claims_exposure = sum_column(rows["claims"], ["claim_amount", "amount", "eot_exposure", "exposure"]) or 0
    status = str(first_valid(rows["projects"], ["status", "project_status"]) or "Active")

    data_quality_fields = [
        contract_value,
        paid_amount,
        planned_progress,
        actual_progress,
        spi,
        cpi,
        risk_score,
    ]
    completeness = sum(value is not None for value in data_quality_fields) / len(data_quality_fields)

    if (spi is not None and spi < 0.9) or (cpi is not None and cpi < 0.9) or high_risk_count > 0:
        decision_required = True
    else:
        decision_required = bool(delay_days or claims_exposure)

    return {
        **{k: v for k, v in project.items() if k != "path"},
        "status": status,
        "contract_value": contract_value,
        "paid_amount": paid_amount,
        "spent_amount": spent_amount,
        "remaining_value": remaining_value,
        "planned_progress": planned_progress,
        "actual_progress": actual_progress,
        "progress_variance": actual_progress - planned_progress if actual_progress is not None and planned_progress is not None else None,
        "bac": bac,
        "pv": pv,
        "ev": ev,
        "ac": ac,
        "spi": spi,
        "cpi": cpi,
        "risk_score": risk_score,
        "high_risk_count": high_risk_count,
        "delay_days": delay_days,
        "claims_exposure": claims_exposure,
        "activity_count": len(rows["activities"]),
        "milestone_count": len(rows["milestones"]),
        "data_quality": round(completeness * 100, 1),
        "decision_required": decision_required,
        "last_updated": latest_mtime(base),
        "fingerprint": fingerprint(base),
        "source_files": {
            key: len(value) for key, value in rows.items()
        },
        "features": build_feature_payload(project, rows),
        "reports": {
            "executive_dashboard": f"/generated/{slugify(project['project_folder_name'])}/01_executive_dashboard.html",
            "master_dashboard": f"/generated/{slugify(project['project_folder_name'])}/02_master_dashboard.html",
            "elite_svg_charts": f"/generated/{slugify(project['project_folder_name'])}/03_elite_svg_charts.html",
            "linked_executive_dashboard": f"/generated/{slugify(project['project_folder_name'])}/04_linked_executive_dashboard.html",
        },
    }


def copy_generated_outputs(projects: list[dict[str, Any]]) -> None:
    GENERATED_ROOT.mkdir(parents=True, exist_ok=True)
    for child in GENERATED_ROOT.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
    for project in projects:
        source = OUTPUTS_ROOT / project["project_folder_name"]
        target = GENERATED_ROOT / slugify(project["project_folder_name"])
        target.mkdir(parents=True, exist_ok=True)
        if not source.exists():
            continue
        for html in sorted(source.glob("*.html")):
            shutil.copy2(html, target / html.name)


def build_portfolio(projects: list[dict[str, Any]]) -> dict[str, Any]:
    total_contract = sum(p["contract_value"] or 0 for p in projects)
    total_paid = sum(p["paid_amount"] or 0 for p in projects)
    total_spent = sum(p["spent_amount"] or 0 for p in projects)
    progress_values = [p["actual_progress"] for p in projects]
    weighted_progress = None
    weighted_basis = sum(p["contract_value"] or 0 for p in projects if p["actual_progress"] is not None)
    if weighted_basis > 0:
        weighted_progress = sum((p["contract_value"] or 0) * (p["actual_progress"] or 0) for p in projects) / weighted_basis

    sectors: dict[str, dict[str, Any]] = {}
    for project in projects:
        sector = sectors.setdefault(
            project["sector"],
            {
                "sector": project["sector"],
                "project_count": 0,
                "contract_value": 0,
                "paid_amount": 0,
                "spent_amount": 0,
                "average_progress": None,
                "average_spi": None,
                "average_cpi": None,
                "average_risk_score": None,
                "delayed_projects": 0,
                "decisions_required": 0,
            },
        )
        sector["project_count"] += 1
        sector["contract_value"] += project["contract_value"] or 0
        sector["paid_amount"] += project["paid_amount"] or 0
        sector["spent_amount"] += project["spent_amount"] or 0
        sector["delayed_projects"] += 1 if (project["delay_days"] or 0) > 0 or (project["spi"] is not None and project["spi"] < 1) else 0
        sector["decisions_required"] += 1 if project["decision_required"] else 0

    for sector in sectors.values():
        sector_projects = [p for p in projects if p["sector"] == sector["sector"]]
        sector["average_progress"] = average([p["actual_progress"] for p in sector_projects])
        sector["average_spi"] = average([p["spi"] for p in sector_projects])
        sector["average_cpi"] = average([p["cpi"] for p in sector_projects])
        sector["average_risk_score"] = average([p["risk_score"] for p in sector_projects])

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "project_count": len(projects),
        "sector_count": len(sectors),
        "totals": {
            "contract_value": total_contract,
            "paid_amount": total_paid,
            "spent_amount": total_spent,
            "remaining_value": total_contract - max(total_paid, total_spent),
            "average_progress": weighted_progress if weighted_progress is not None else average(progress_values),
            "average_spi": average([p["spi"] for p in projects]),
            "average_cpi": average([p["cpi"] for p in projects]),
            "average_risk_score": average([p["risk_score"] for p in projects]),
            "delayed_projects": sum(1 for p in projects if (p["delay_days"] or 0) > 0 or (p["spi"] is not None and p["spi"] < 1)),
            "high_risk_projects": sum(1 for p in projects if (p["risk_score"] or 0) >= 70 or p["high_risk_count"] > 0),
            "claims_exposure": sum(p["claims_exposure"] or 0 for p in projects),
            "decisions_required": sum(1 for p in projects if p["decision_required"]),
        },
        "sectors": sorted(sectors.values(), key=lambda item: item["sector"]),
        "projects": sorted(projects, key=lambda item: (item["sector"], item["project_display_name"])),
    }


def main() -> None:
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    raw_projects = discover_projects()
    project_records = [build_project_record(project) for project in raw_projects]
    copy_generated_outputs(project_records)
    portfolio = build_portfolio(project_records)
    (DATA_ROOT / "portfolio.json").write_text(json.dumps(portfolio, indent=2, ensure_ascii=False), encoding="utf-8")
    projects_dir = DATA_ROOT / "projects"
    projects_dir.mkdir(parents=True, exist_ok=True)
    for stale in projects_dir.glob("*.json"):
        stale.unlink()
    for project in project_records:
        (projects_dir / f"{project['project_key']}.json").write_text(json.dumps(project, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Generated Next.js website data for {len(project_records)} projects.")


if __name__ == "__main__":
    main()
